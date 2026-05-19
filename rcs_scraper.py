#!/usr/bin/env python3
"""
Scraper RCS Madagascar (https://www.rcsmada.mg)

Installation des dépendances :
    pip install -r requirements.txt

Lancer sans argument pour ouvrir l'interface graphique :
    python3 rcs_scraper.py

Ou en ligne de commande :
    python3 rcs_scraper.py --greffe "Nosy Be" --annee 2025 --forme SARL
    python3 rcs_scraper.py --greffe "Antananarivo" --annee 2026 --forme SA-AG --type B
"""

import argparse
import html
import os
import queue
import re
import subprocess
import sys
import threading
import time
from datetime import datetime
from pathlib import Path

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    _TK_IMPORT_ERROR = None
except ImportError as _tk_err:
    tk = None  # type: ignore
    filedialog = None  # type: ignore
    messagebox = None  # type: ignore
    ttk = None  # type: ignore
    _TK_IMPORT_ERROR = str(_tk_err)

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://www.rcsmada.mg"
LIST_URL = f"{BASE_URL}/index.php?pgdown=liste2"
DETAIL_URL = f"{BASE_URL}/index.php?pgdown=consultation&soc={{soc}}"
REFERER = f"{BASE_URL}/index.php?pgdown=liste&pgmenu=Services%20Propos%C3%A9s"

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)

GREFFE_MAP = {
    "Ambositra": "1", "Antananarivo": "10", "Antsirabe": "11",
    "Antsiranana": "12", "Arivonimamo": "13", "Betroka": "14",
    "Farafangana": "15", "Fianarantsoa": "16", "Ihosy": "17",
    "Maevatanana": "18", "Mahajanga": "19", "Ampanihy": "2",
    "Maintirano": "20", "Mampikony": "21", "Manakara": "22",
    "Mananjary": "23", "Mandritsara": "24", "Maroantsetra": "25",
    "Miandrivazo": "26", "Miarinarivo": "27", "Moramanga": "28",
    "Morombe": "29", "Ambanja": "3", "Morondava": "30",
    "Nosy Be": "31", "Port Berger": "32", "Taolagnaro": "33",
    "Toliara": "34", "Vatomandry": "35", "Toamasina": "36",
    "Antsohihy": "37", "Ikongo": "38", "Tsiroanomandidy": "39",
    "Ambatolampy": "4", "Ambovombe": "40", "Avaradrano": "43",
    "Ambatondrazaka": "5", "Sainte Marie": "51", "Ambilobe": "52",
    "Analalava": "6", "Ankazoabo": "7", "Ankazobe": "8",
    "Antalaha": "9",
}

TYPE_LABELS = {
    "A": "A - Personne physique",
    "B": "B - Personne morale",
    "C": "C - Groupement d'intérêt économique",
    "D": "D - Personne morale autre qu'un GIE",
    "E": "E - Institution de microfinance",
}

# Codes officiels acceptés par https://www.rcsmada.mg (champ "FormeJuridiq" du formulaire).
# Ordre = utilisation décroissante. Clé = code envoyé au serveur, valeur = libellé lisible.
FORMES_LABELS = {
    "SARL": "Société à Responsabilité Limitée",
    "SRL U": "SARL Unipersonnelle",
    "SA-AG": "Société Anonyme avec Administrateur Général",
    "SA-CA": "Société Anonyme avec Conseil d'Administration",
    "GIE": "Groupement d'Intérêt Économique",
    "ASS": "Association",
    "COOP": "Coopérative",
    "SCI": "Société Civile",
    "SCS": "Société en Commandite Simple",
    "SCA": "Société en Commandite par Action",
    "SNC": "Société en Nom Commercial",
    "SDET": "Société de Droit Étranger",
    "EPIC": "Établissement Public à Caractère Industriel et Commercial",
    "IMF": "Institut de Micro Finance",
    "IMF-SA": "IMF - SA",
    "IMF-SARL": "IMF - SARL",
    "IMF-ASS": "IMF - Association",
    "IMF-COOP": "IMF - Coopérative",
    "AUTRES": "Autres",
}
FORMES_COURANTES = list(FORMES_LABELS.keys())

_MOJIBAKE_FIXES = [
    ("Ã©", "é"), ("Ã¨", "è"), ("Ãª", "ê"), ("Ã«", "ë"),
    ("Ã ", "à"), ("Ã¢", "â"), ("Ã¤", "ä"), ("Ã¥", "å"),
    ("Ã®", "î"), ("Ã¯", "ï"),
    ("Ã´", "ô"), ("Ã¶", "ö"), ("Ã²", "ò"), ("Ã³", "ó"),
    ("Ã¹", "ù"), ("Ã»", "û"), ("Ã¼", "ü"),
    ("Ã§", "ç"), ("Ã±", "ñ"),
    ("Ã‰", "É"), ("Ãˆ", "È"), ("ÃŠ", "Ê"),
    ("Ã€", "À"), ("Ã‚", "Â"),
    ('Ã"', "Ô"), ("Ã–", "Ö"),
    ("Ã™", "Ù"), ("Ã›", "Û"), ("Ãœ", "Ü"),
    ("Ã‡", "Ç"),
    ("Ã¦", "æ"), ("Ã†", "Æ"),
    ('Å"', "œ"), ("Å'", "Œ"),
    ("Â°", "°"), ("Â«", "«"), ("Â»", "»"),
]

_HTML_ENTITIES = "|".join([
    "eacute", "egrave", "ecirc", "euml",
    "Eacute", "Egrave", "Ecirc", "Euml",
    "agrave", "acirc", "auml", "atilde", "aelig", "aring",
    "Agrave", "Acirc", "Auml", "Atilde", "Aelig", "Aring",
    "iacute", "igrave", "icirc", "iuml",
    "Iacute", "Igrave", "Icirc", "Iuml",
    "oacute", "ograve", "ocirc", "ouml", "otilde", "oelig",
    "Oacute", "Ograve", "Ocirc", "Ouml", "Otilde", "Oelig",
    "uacute", "ugrave", "ucirc", "uuml",
    "Uacute", "Ugrave", "Ucirc", "Uuml",
    "ccedil", "Ccedil", "ntilde", "Ntilde",
    "yacute", "yuml", "Yacute", "Yuml",
    "nbsp", "amp", "quot", "apos", "lt", "gt", "deg", "copy", "reg", "euro",
])

EXCEL_HEADERS = [
    "Immatriculation",
    "Dénomination",
    "Adresse",
    "Date de début",
    "Date d'immatriculation",
    "Capital",
    "Forme juridique",
    "Activité",
    "Type d'entreprise",
    "Dirigeant - Fonction",
    "Dirigeant - Nom et prénoms",
]


def make_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Referer": REFERER,
        "Accept-Language": "fr-FR,fr;q=0.9",
    })
    return s


def request_with_retry(session, method, url, **kwargs):
    """GET/POST avec retry et backoff exponentiel sur 403/429/5xx/timeout."""
    max_attempts = 5
    delay = 2.0
    last_err = None
    for attempt in range(1, max_attempts + 1):
        try:
            r = session.request(method, url, timeout=30, **kwargs)
            if r.status_code in (403, 429) or r.status_code >= 500:
                last_err = f"HTTP {r.status_code}"
            else:
                return r
        except requests.RequestException as e:
            last_err = str(e)
        if attempt < max_attempts:
            print(f"    !! {last_err} - retry dans {delay:.0f}s (essai {attempt}/{max_attempts})", file=sys.stderr)
            time.sleep(delay)
            delay *= 2
    raise RuntimeError(f"Échec après {max_attempts} essais: {last_err}")


def fetch_list_page1(session, type_societe, greffe_id, annee, forme):
    payload = {
        "TypeSociete": type_societe,
        "Greffe": str(greffe_id),
        "DateInscrit": str(annee),
        "FormeJuridiq": forme,
    }
    r = request_with_retry(session, "POST", LIST_URL, data=payload)
    r.encoding = r.encoding or "utf-8"
    return r.text


def fetch_list_page_n(session, page, req):
    url = f"{BASE_URL}/index.php"
    params = {"pgdown": "liste2", "page": str(page), "req": req}
    r = request_with_retry(session, "GET", url, params=params)
    r.encoding = r.encoding or "utf-8"
    return r.text


def parse_pagination(html):
    """Extrait le nombre total de pages et le paramètre 'req' (encodé) pour la pagination."""
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(" ", strip=True)
    m = re.search(r"Nombre de page\s*:\s*(\d+)", text)
    nb_pages = int(m.group(1)) if m else 1
    req = None
    for a in soup.select('a[href*="page="]'):
        href = a.get("href", "")
        m2 = re.search(r"req=([^&\s\"']+)", href)
        if m2:
            req = m2.group(1)
            break
    return nb_pages, req


def fetch_all_pages(session, type_societe, greffe_id, annee, forme, sleep_s=1.0, progress_cb=None, quiet=False):
    """Récupère TOUTES les pages de résultats et concatène les entrées."""
    if not quiet:
        print("-> POST sur le formulaire (page 1)...")
    if progress_cb:
        progress_cb("status", message="Recherche en cours (page 1)...")
    html_p1 = fetch_list_page1(session, type_societe, greffe_id, annee, forme)
    nb_pages, req = parse_pagination(html_p1)
    if not quiet:
        print(f"-> Pagination détectée: {nb_pages} page(s)")
    if progress_cb:
        progress_cb("status", message=f"{nb_pages} page(s) de résultats détectée(s)...")

    all_entries = parse_list(html_p1)

    if nb_pages > 1 and req:
        for page in range(2, nb_pages + 1):
            time.sleep(sleep_s)
            if not quiet:
                print(f"-> Récupération de la page {page}/{nb_pages}...")
            if progress_cb:
                progress_cb("status", message=f"Récupération de la liste (page {page}/{nb_pages})...")
            html_pn = fetch_list_page_n(session, page, req)
            entries_n = parse_list(html_pn)
            all_entries.extend(entries_n)

    seen = set()
    deduped = []
    for e in all_entries:
        if e["soc_id"] in seen:
            continue
        seen.add(e["soc_id"])
        deduped.append(e)
    return deduped


def parse_list(html):
    soup = BeautifulSoup(html, "html.parser")
    seen = set()
    entries = []
    for a in soup.select('a[href*="pgdown=consultation"]'):
        href = a.get("href", "")
        m = re.search(r"soc=([\w-]+)", href)
        if not m:
            continue
        soc_id = m.group(1)
        if soc_id in seen:
            continue
        seen.add(soc_id)
        tr = a.find_parent("tr")
        immat = denom = ""
        if tr:
            tds = tr.find_all("td")
            if len(tds) >= 1:
                immat = tds[0].get_text(" ", strip=True)
            if len(tds) >= 2:
                denom = tds[1].get_text(" ", strip=True)
        entries.append({
            "soc_id": soc_id,
            "immatriculation_liste": immat,
            "denomination_liste": denom,
        })
    return entries


def clean_value(text):
    if text is None:
        return ""
    text = text.strip()
    if text.startswith(":"):
        text = text[1:].strip()
    # Le serveur renvoie des entités HTML mal formées sans ";" (ex: "&eacute" au lieu de "&eacute;").
    # On répare en se basant sur une liste explicite d'entités valides pour éviter de matcher trop large.
    text = re.sub(rf"&({_HTML_ENTITIES})(?!;)", r"&\1;", text)
    text = html.unescape(text)
    text = text.replace("''", "'")
    # Corrige les séquences mojibake UTF-8/Latin-1 résiduelles dans la base source.
    for bad, good in _MOJIBAKE_FIXES:
        text = text.replace(bad, good)
    return re.sub(r"\s+", " ", text)


def parse_detail(html):
    """Extrait les infos de la page consultation. Robuste aux variations de mise en page."""
    soup = BeautifulSoup(html, "html.parser")
    data = {h: "" for h in EXCEL_HEADERS}

    table = soup.find("table", class_="simple2")
    if not table:
        return data

    # Ordre important: needles plus spécifiques en premier pour éviter qu'un préfixe
    # commun ne matche par erreur (ex: "immatriculation" matchait "date d'immatriculation").
    label_to_key = [
        ("date d'immatriculation", "Date d'immatriculation"),
        ("date de début", "Date de début"),
        ("type d'entreprise", "Type d'entreprise"),
        ("forme juridique", "Forme juridique"),
        ("dénomination", "Dénomination"),
        ("denomination", "Dénomination"),
        ("immatriculation", "Immatriculation"),
        ("adresse", "Adresse"),
        ("capital", "Capital"),
        ("activité", "Activité"),
        ("activite", "Activité"),
    ]

    for tr in table.find_all("tr", recursive=False):
        tds = tr.find_all("td", recursive=False)
        if len(tds) < 2:
            continue
        label = tds[0].get_text(" ", strip=True).lower().rstrip(":").strip()

        if "dirigeant" in label:
            inner = tds[1].find("table")
            if inner:
                rows = inner.find_all("tr")
                if rows:
                    fonction = rows[0].get_text(" ", strip=True)
                    data["Dirigeant - Fonction"] = fonction
                for r in rows[1:]:
                    cells = r.find_all("td")
                    if len(cells) >= 2:
                        cl = cells[0].get_text(" ", strip=True).lower()
                        if "nom" in cl:
                            data["Dirigeant - Nom et prénoms"] = clean_value(cells[1].get_text(" ", strip=True))
            continue

        for needle, key in label_to_key:
            if needle in label:
                value = clean_value(tds[1].get_text(" ", strip=True))
                data[key] = value
                break

    return data


def fetch_detail(session, soc_id):
    url = DETAIL_URL.format(soc=soc_id)
    r = request_with_retry(session, "GET", url)
    r.encoding = r.encoding or "utf-8"
    return r.text


def style_workbook(ws, row_count):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    zebra_fill = PatternFill("solid", fgColor="F2F6FA")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_idx, _ in enumerate(EXCEL_HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 32

    for row in range(2, row_count + 2):
        for col in range(1, len(EXCEL_HEADERS) + 1):
            c = ws.cell(row=row, column=col)
            c.alignment = left_wrap
            c.border = border
            c.font = Font(name="Calibri", size=10)
            if row % 2 == 0:
                c.fill = zebra_fill

    widths = {
        "Immatriculation": 26,
        "Dénomination": 32,
        "Adresse": 45,
        "Date de début": 15,
        "Date d'immatriculation": 18,
        "Capital": 14,
        "Forme juridique": 32,
        "Activité": 45,
        "Type d'entreprise": 20,
        "Dirigeant - Fonction": 16,
        "Dirigeant - Nom et prénoms": 28,
    }
    for col_idx, h in enumerate(EXCEL_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 18)

    ws.freeze_panes = "A2"
    if row_count > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_HEADERS))}{row_count + 1}"


def build_workbook(rows, params):
    wb = Workbook()
    ws = wb.active
    ws.title = params["greffe"][:31]

    ws.append(EXCEL_HEADERS)
    for r in rows:
        ws.append([r.get(h, "") for h in EXCEL_HEADERS])
    style_workbook(ws, len(rows))

    meta = wb.create_sheet("Info")
    meta.append(["Paramètres de recherche", ""])
    meta.append(["Greffe", params["greffe"]])
    meta.append(["Année", params["annee"]])
    meta.append(["Forme juridique", params["forme"]])
    meta.append(["Type d'assujetti", TYPE_LABELS.get(params["type"], params["type"])])
    meta.append(["Nombre d'entreprises", len(rows)])
    meta.append(["Date d'extraction", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["Source", BASE_URL])
    meta["A1"].font = Font(bold=True, size=12)
    meta.column_dimensions["A"].width = 26
    meta.column_dimensions["B"].width = 40
    for row in range(2, 9):
        meta[f"A{row}"].font = Font(bold=True)
    return wb


def extract_entreprises(session, type_code, greffe_id, annee, forme, sleep_s=1.0, limit=0, progress_cb=None, quiet=False):
    """Récupère la liste + le détail de chaque entreprise. Retourne une liste de dict."""
    entries = fetch_all_pages(session, type_code, greffe_id, annee, forme,
                              sleep_s=sleep_s, progress_cb=progress_cb, quiet=quiet)
    if not quiet:
        print(f"-> {len(entries)} entreprises trouvées au total.")

    if limit > 0:
        entries = entries[:limit]
        if not quiet:
            print(f"-> Limite appliquée: {len(entries)} entreprises seront traitées.")

    total = len(entries)
    if progress_cb:
        progress_cb("progress", current=0, total=total, eta=total * (sleep_s + 1.0),
                    message=f"{total} entreprise(s) trouvée(s). Récupération des détails...")

    rows = []
    start = time.time()
    for i, e in enumerate(entries, 1):
        time.sleep(sleep_s)
        if not quiet:
            print(f"  [{i}/{total}] {e['immatriculation_liste']} - {e['denomination_liste']}")
        try:
            html_detail = fetch_detail(session, e["soc_id"])
            detail = parse_detail(html_detail)
        except Exception as ex:
            if not quiet:
                print(f"    !! erreur sur {e['soc_id']}: {ex}", file=sys.stderr)
            detail = {h: "" for h in EXCEL_HEADERS}
            detail["Immatriculation"] = e["immatriculation_liste"]
            detail["Dénomination"] = e["denomination_liste"]

        if not detail.get("Immatriculation"):
            detail["Immatriculation"] = e["immatriculation_liste"]
        if not detail.get("Dénomination"):
            detail["Dénomination"] = e["denomination_liste"]
        rows.append(detail)

        if progress_cb:
            elapsed = time.time() - start
            avg = elapsed / i
            remaining = avg * (total - i)
            label = e["denomination_liste"] or e["immatriculation_liste"]
            progress_cb("progress", current=i, total=total, eta=remaining,
                        message=f"[{i}/{total}] {label[:60]}")

    return rows


def format_duration(seconds):
    seconds = max(0, int(seconds))
    if seconds < 60:
        return f"{seconds} s"
    m, s = divmod(seconds, 60)
    if m < 60:
        return f"{m} min {s:02d} s"
    h, m = divmod(m, 60)
    return f"{h} h {m:02d} min"


def _likely_broken_macos_tk():
    """Détecte la combinaison Python 3.9 système macOS, connue pour avoir un Tk
    compilé contre un SDK macOS plus récent que le système hôte (le tk.Tk()
    appelle alors Tcl_Panic() puis abort())."""
    if sys.platform != "darwin":
        return False
    if sys.version_info[:2] != (3, 9):
        return False
    exe = sys.executable or ""
    return exe.startswith("/usr/bin/") or "/Library/Developer/CommandLineTools/" in exe


def can_use_gui():
    """Teste si Tk peut réellement s'initialiser, sans risquer d'aborter le
    processus appelant. Sur certaines installations (Python 3.9 système macOS),
    tk.Tk() appelle Tcl_Panic() puis abort() — non rattrapable par try/except."""
    if tk is None:
        return False
    # En mode PyInstaller bundle, Tk est embarqué et sys.executable pointe sur
    # l'app elle-même : relancer un sous-processus ferait re-démarrer l'app.
    if getattr(sys, "frozen", False):
        return True
    if _likely_broken_macos_tk():
        return False
    try:
        result = subprocess.run(
            [sys.executable, "-c",
             "import tkinter; r = tkinter.Tk(); r.withdraw(); r.destroy()"],
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=10,
        )
        return result.returncode == 0
    except Exception:
        return False


_GUI_INSTALL_HINTS = {
    "darwin": (
        "  - macOS : installez Python depuis https://www.python.org/downloads/\n"
        "             (qui inclut un Tk compatible)\n"
        "             ou via Homebrew : brew install python-tk@3.12"
    ),
    "linux": (
        "  - Debian/Ubuntu : sudo apt install python3-tk\n"
        "  - Fedora        : sudo dnf install python3-tkinter\n"
        "  - Arch          : sudo pacman -S tk"
    ),
    "win": (
        "  - Windows : réinstallez Python depuis https://www.python.org/downloads/\n"
        "              en cochant l'option « tcl/tk and IDLE »"
    ),
}


def _gui_install_hint():
    """Conseils pour activer l'interface graphique selon la plateforme."""
    platform = sys.platform
    for prefix, hint in _GUI_INSTALL_HINTS.items():
        if platform == prefix or platform.startswith(prefix):
            return hint
    return ""


def _print_progress_bar(current, total, eta_seconds=0, message=""):
    bar_len = 30
    if total > 0:
        pct = current / total * 100
        filled = int(bar_len * current / total)
    else:
        pct = 0
        filled = 0
    bar = "█" * filled + "░" * (bar_len - filled)
    eta_str = f" ~{format_duration(eta_seconds)} restant" if eta_seconds > 0 else ""
    msg = message or ""
    if len(msg) > 50:
        msg = msg[:47] + "..."
    line = f"\r[{bar}] {pct:3.0f}% ({current}/{total}){eta_str}  {msg}"
    sys.stdout.write(line.ljust(110))
    sys.stdout.flush()


def _prompt(question, default=None):
    suffix = f" [{default}]" if default is not None else ""
    while True:
        try:
            val = input(f"{question}{suffix} : ").strip()
        except EOFError:
            print()
            sys.exit(0)
        if not val and default is not None:
            return default
        if val:
            return val
        print("  ! Réponse requise.")


def _prompt_choice(question, options, default_key):
    """options : liste de (key, label). Retourne la clé choisie."""
    print(f"\n{question}")
    for i, (key, label) in enumerate(options, 1):
        marker = "  (défaut)" if key == default_key else ""
        print(f"  {i}) {label}{marker}")
    default_idx = next((i for i, (k, _) in enumerate(options, 1) if k == default_key), 1)
    while True:
        try:
            val = input(f"Choix [{default_idx}] : ").strip()
        except EOFError:
            print()
            sys.exit(0)
        if not val:
            return options[default_idx - 1][0]
        if val.isdigit():
            idx = int(val)
            if 1 <= idx <= len(options):
                return options[idx - 1][0]
        for k, _ in options:
            if val.upper() == k.upper():
                return k
        print(f"  ! Entrez un nombre entre 1 et {len(options)} (ou la lettre A/B/C/D/E).")


def _prompt_greffe(default="Nosy Be"):
    sorted_greffes = sorted(GREFFE_MAP.keys())
    print("\nListe des greffes disponibles :")
    col_width = 22
    cols = 3
    for i in range(0, len(sorted_greffes), cols):
        row = sorted_greffes[i:i + cols]
        print("  " + "".join(g.ljust(col_width) for g in row))
    while True:
        val = _prompt("\nVille (greffe)", default)
        if val in GREFFE_MAP:
            return val
        for g in GREFFE_MAP:
            if g.lower() == val.lower():
                return g
        suggestions = [g for g in GREFFE_MAP if val.lower() in g.lower()][:5]
        if suggestions:
            print(f"  ! Greffe inconnu. Suggestions : {', '.join(suggestions)}")
        else:
            print(f"  ! Greffe inconnu : {val!r}")


def launch_interactive(reason=None):
    """Mode interactif texte (fallback quand l'interface graphique est indisponible)."""
    print()
    print("=" * 64)
    print("  RCS Madagascar - Extracteur d'entreprises")
    print("=" * 64)
    if reason:
        print()
        print(f"  Note : {reason}")
        hint = _gui_install_hint()
        if hint:
            print("  Pour activer l'interface graphique :")
            print(hint)
    print()
    print("Mode interactif texte. Appuyez sur Entrée pour accepter la valeur par défaut.")
    print("-" * 64)

    greffe = _prompt_greffe(default="Nosy Be")

    while True:
        annee_str = _prompt("Année", default=str(datetime.now().year))
        try:
            annee = int(annee_str)
            break
        except ValueError:
            print("  ! L'année doit être un nombre entier.")

    print(f"\nFormes courantes : {', '.join(FORMES_COURANTES)}")
    forme = _prompt("Forme juridique", default="SARL")

    type_options = list(TYPE_LABELS.items())
    type_code = _prompt_choice("Type d'assujetti :", type_options, default_key="B")

    while True:
        output_dir = _prompt("Dossier de sortie", default=str(Path.home() / "Desktop"))
        out_dir = Path(output_dir).expanduser()
        try:
            out_dir.mkdir(parents=True, exist_ok=True)
            break
        except Exception as e:
            print(f"  ! Dossier invalide : {e}")

    output_file = out_dir / f"rcs_{greffe.replace(' ', '_')}_{annee}_{forme}.xlsx"

    print()
    print("-" * 64)
    print("Récapitulatif :")
    print(f"  Ville    : {greffe}")
    print(f"  Année    : {annee}")
    print(f"  Forme    : {forme}")
    print(f"  Type     : {TYPE_LABELS[type_code]}")
    print(f"  Sortie   : {output_file}")
    print("-" * 64)

    try:
        confirm = input("\nLancer l'extraction ? [O/n] : ").strip().lower()
    except EOFError:
        confirm = "o"
    if confirm and confirm not in ("o", "oui", "y", "yes"):
        print("Annulé.")
        return

    print("\nLancement de l'extraction...\n")
    start = time.time()
    state = {"last_msg": "Connexion au site..."}

    def cb(event, **kwargs):
        if event == "status":
            state["last_msg"] = kwargs.get("message", state["last_msg"])
            _print_progress_bar(0, 0, 0, state["last_msg"])
        elif event == "progress":
            state["last_msg"] = kwargs.get("message", state["last_msg"])
            _print_progress_bar(
                kwargs.get("current", 0),
                kwargs.get("total", 0),
                kwargs.get("eta", 0),
                state["last_msg"],
            )

    try:
        session = make_session()
        rows = extract_entreprises(session, type_code, GREFFE_MAP[greffe], str(annee), forme,
                                   sleep_s=1.0, progress_cb=cb, quiet=True)
        _print_progress_bar(len(rows), max(len(rows), 1), 0, "Génération du fichier Excel...")
        wb = build_workbook(rows, {
            "greffe": greffe, "annee": str(annee), "forme": forme, "type": type_code,
        })
        wb.save(str(output_file))
        elapsed = time.time() - start
        print("\n")
        print("=" * 64)
        print(f"  Extraction terminée en {format_duration(elapsed)}")
        print(f"  {len(rows)} entreprise(s) extraite(s)")
        print(f"  Fichier : {output_file}")
        print("=" * 64)
    except KeyboardInterrupt:
        print("\n\nInterrompu par l'utilisateur.")
        sys.exit(130)
    except Exception as e:
        print("\n")
        print(f"Erreur : {e}", file=sys.stderr)
        sys.exit(1)


class ScraperGUI:
    PRIMARY = "#1F4E78"
    PRIMARY_LIGHT = "#2E75B6"
    BG = "#F7F9FC"
    SURFACE = "#FFFFFF"
    BORDER = "#D9E2EC"
    MUTED = "#5A6B7B"
    TEXT = "#1A2733"
    SUCCESS = "#2E7D32"
    ERROR = "#C62828"

    def __init__(self, root):
        self.root = root
        self.root.title("RCS Madagascar — Extracteur d'entreprises")
        self.root.geometry("760x780")
        self.root.minsize(720, 740)
        self.root.configure(bg=self.BG)

        self.queue = queue.Queue()
        self.worker = None
        self.is_running = False
        self.start_time = None
        self.output_path = None

        self._setup_styles()
        self._build_ui()
        self._poll_queue()

    def _setup_styles(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(".", background=self.BG, foreground=self.TEXT, font=("Helvetica", 11))
        style.configure("TFrame", background=self.BG)
        style.configure("Surface.TFrame", background=self.SURFACE)
        style.configure("TLabel", background=self.BG, foreground=self.TEXT)
        style.configure("Surface.TLabel", background=self.SURFACE, foreground=self.TEXT)
        style.configure("Title.TLabel", background=self.BG, foreground=self.PRIMARY,
                        font=("Helvetica", 20, "bold"))
        style.configure("Subtitle.TLabel", background=self.BG, foreground=self.MUTED,
                        font=("Helvetica", 11))
        style.configure("SectionTitle.TLabel", background=self.SURFACE, foreground=self.PRIMARY,
                        font=("Helvetica", 12, "bold"))
        style.configure("FieldLabel.TLabel", background=self.SURFACE, foreground=self.TEXT,
                        font=("Helvetica", 10, "bold"))
        style.configure("FieldHelp.TLabel", background=self.SURFACE, foreground=self.MUTED,
                        font=("Helvetica", 9))
        style.configure("Status.TLabel", background=self.SURFACE, foreground=self.TEXT,
                        font=("Helvetica", 10))
        style.configure("Percent.TLabel", background=self.SURFACE, foreground=self.PRIMARY,
                        font=("Helvetica", 14, "bold"))
        style.configure("Eta.TLabel", background=self.SURFACE, foreground=self.MUTED,
                        font=("Helvetica", 10))
        style.configure("Success.TLabel", background=self.SURFACE, foreground=self.SUCCESS,
                        font=("Helvetica", 11, "bold"))
        style.configure("Error.TLabel", background=self.SURFACE, foreground=self.ERROR,
                        font=("Helvetica", 11, "bold"))

        style.configure("Primary.TButton",
                        background=self.PRIMARY, foreground="#FFFFFF",
                        font=("Helvetica", 12, "bold"), padding=(20, 12),
                        borderwidth=0, focusthickness=0)
        style.map("Primary.TButton",
                  background=[("active", self.PRIMARY_LIGHT), ("disabled", "#9AB3C9")],
                  foreground=[("disabled", "#E8EFF7")])

        style.configure("Secondary.TButton",
                        background=self.SURFACE, foreground=self.PRIMARY,
                        font=("Helvetica", 10), padding=(12, 6),
                        borderwidth=1, relief="solid")
        style.map("Secondary.TButton",
                  background=[("active", "#EDF3FA")])

        style.configure("TEntry", fieldbackground="#FFFFFF", borderwidth=1, padding=6)
        style.configure("TCombobox", fieldbackground="#FFFFFF", background="#FFFFFF",
                        arrowcolor=self.PRIMARY, padding=4)
        style.configure("TSpinbox", fieldbackground="#FFFFFF", arrowcolor=self.PRIMARY, padding=4)
        style.map("TCombobox",
                  fieldbackground=[("readonly", "#FFFFFF")],
                  selectbackground=[("readonly", "#FFFFFF")],
                  selectforeground=[("readonly", self.TEXT)])

        style.configure("Blue.Horizontal.TProgressbar",
                        troughcolor="#E8EFF7", background=self.PRIMARY,
                        thickness=18, borderwidth=0)

    def _card(self, parent):
        outer = tk.Frame(parent, bg=self.BORDER, bd=0)
        inner = tk.Frame(outer, bg=self.SURFACE)
        inner.pack(fill="both", expand=True, padx=1, pady=1)
        return outer, inner

    def _build_ui(self):
        wrapper = ttk.Frame(self.root, padding=(24, 20, 24, 20))
        wrapper.pack(fill="both", expand=True)

        header = ttk.Frame(wrapper)
        header.pack(fill="x", pady=(0, 18))
        ttk.Label(header, text="RCS Madagascar", style="Title.TLabel").pack(anchor="w")
        ttk.Label(header,
                  text="Extracteur d'entreprises du Registre du Commerce et des Sociétés",
                  style="Subtitle.TLabel").pack(anchor="w", pady=(4, 0))

        form_outer, form = self._card(wrapper)
        form_outer.pack(fill="x", pady=(0, 14))
        form_inner = tk.Frame(form, bg=self.SURFACE, padx=22, pady=20)
        form_inner.pack(fill="both", expand=True)
        ttk.Label(form_inner, text="Critères de recherche",
                  style="SectionTitle.TLabel").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 14))
        form_inner.columnconfigure(0, weight=1, uniform="col")
        form_inner.columnconfigure(1, weight=1, uniform="col")

        self.var_greffe = tk.StringVar(value="Nosy Be")
        greffe_cell = self._field(form_inner, 1, 0, "Ville (greffe)",
                                  "La ville où l'entreprise est immatriculée")
        ttk.Combobox(greffe_cell, textvariable=self.var_greffe,
                     values=sorted(GREFFE_MAP.keys()),
                     state="readonly").grid(row=2, column=0, sticky="ew")

        self.var_annee = tk.StringVar(value=str(datetime.now().year))
        annee_cell = self._field(form_inner, 1, 1, "Année",
                                 "Année d'immatriculation au RCS")
        ttk.Spinbox(annee_cell, from_=2000, to=datetime.now().year + 1,
                    textvariable=self.var_annee,
                    justify="left").grid(row=2, column=0, sticky="ew")

        self.var_forme = tk.StringVar(value="SARL")
        forme_cell = self._field(form_inner, 3, 0, "Forme juridique",
                                 "Exemples : SARL, SA-AG, EI, GIE…")
        ttk.Combobox(forme_cell, textvariable=self.var_forme,
                     values=FORMES_COURANTES).grid(row=2, column=0, sticky="ew")

        self.var_type = tk.StringVar(value=TYPE_LABELS["B"])
        type_cell = self._field(form_inner, 3, 1, "Type d'assujetti",
                                "Personne physique, morale, GIE, etc.")
        ttk.Combobox(type_cell, textvariable=self.var_type,
                     values=list(TYPE_LABELS.values()),
                     state="readonly").grid(row=2, column=0, sticky="ew")

        out_outer, out_card = self._card(wrapper)
        out_outer.pack(fill="x", pady=(0, 14))
        out_inner = tk.Frame(out_card, bg=self.SURFACE, padx=22, pady=20)
        out_inner.pack(fill="both", expand=True)
        out_inner.columnconfigure(0, weight=1)
        ttk.Label(out_inner, text="Fichier de sortie",
                  style="SectionTitle.TLabel").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 12))
        ttk.Label(out_inner, text="Dossier de destination",
                  style="FieldLabel.TLabel").grid(row=1, column=0, columnspan=2, sticky="w")
        ttk.Label(out_inner, text="Le fichier Excel sera créé dans ce dossier",
                  style="FieldHelp.TLabel").grid(row=2, column=0, columnspan=2, sticky="w", pady=(2, 6))

        self.var_output = tk.StringVar(value=str(Path.home() / "Desktop"))
        ttk.Entry(out_inner, textvariable=self.var_output).grid(row=3, column=0, sticky="ew", padx=(0, 8))
        ttk.Button(out_inner, text="Parcourir…", command=self._browse,
                   style="Secondary.TButton").grid(row=3, column=1)

        self.btn_run = ttk.Button(wrapper, text="Lancer l'extraction",
                                  command=self._start, style="Primary.TButton")
        self.btn_run.pack(fill="x", pady=(0, 14))

        progress_outer, progress_card = self._card(wrapper)
        progress_outer.pack(fill="both", expand=True)
        progress_inner = tk.Frame(progress_card, bg=self.SURFACE, padx=22, pady=20)
        progress_inner.pack(fill="both", expand=True)
        progress_inner.columnconfigure(0, weight=1)

        ttk.Label(progress_inner, text="Progression",
                  style="SectionTitle.TLabel").grid(row=0, column=0, columnspan=2, sticky="w")

        self.var_status = tk.StringVar(value="Prêt. Configurez vos critères puis cliquez sur « Lancer l'extraction ».")
        self.lbl_status = ttk.Label(progress_inner, textvariable=self.var_status,
                                    style="Status.TLabel", wraplength=640, justify="left")
        self.lbl_status.grid(row=1, column=0, columnspan=2, sticky="w", pady=(10, 12))

        self.progress = ttk.Progressbar(progress_inner, mode="determinate",
                                        style="Blue.Horizontal.TProgressbar", maximum=100)
        self.progress.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(0, 10))

        self.var_percent = tk.StringVar(value="0 %")
        ttk.Label(progress_inner, textvariable=self.var_percent,
                  style="Percent.TLabel").grid(row=3, column=0, sticky="w")
        self.var_eta = tk.StringVar(value="Aucune extraction en cours")
        ttk.Label(progress_inner, textvariable=self.var_eta,
                  style="Eta.TLabel").grid(row=3, column=1, sticky="e")

        self.var_result = tk.StringVar(value="")
        self.lbl_result = ttk.Label(progress_inner, textvariable=self.var_result,
                                    style="Success.TLabel", wraplength=640, justify="left")
        self.lbl_result.grid(row=4, column=0, columnspan=2, sticky="w", pady=(14, 0))

        self.btn_open = ttk.Button(progress_inner, text="Ouvrir le dossier",
                                   command=self._open_output_folder, style="Secondary.TButton")

    def _field(self, parent, row, col, label, help_text):
        """Crée une cellule de formulaire (label + help text) et retourne le frame.
        L'appelant y place le widget de saisie en row=2, column=0."""
        pad_x = (0, 12) if col == 0 else (12, 0)
        cell = tk.Frame(parent, bg=self.SURFACE)
        cell.grid(row=row, column=col, sticky="ew", pady=(0, 14), padx=pad_x)
        cell.columnconfigure(0, weight=1)
        ttk.Label(cell, text=label, style="FieldLabel.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(cell, text=help_text, style="FieldHelp.TLabel").grid(row=1, column=0, sticky="w", pady=(1, 6))
        return cell

    def _browse(self):
        path = filedialog.askdirectory(initialdir=self.var_output.get() or str(Path.home()),
                                       title="Choisir le dossier de destination")
        if path:
            self.var_output.set(path)

    def _open_output_folder(self):
        if not self.output_path:
            return
        folder = str(Path(self.output_path).parent)
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            elif sys.platform.startswith("win"):
                os.startfile(folder)
            else:
                subprocess.Popen(["xdg-open", folder])
        except Exception as e:
            messagebox.showwarning("Information", f"Impossible d'ouvrir le dossier :\n{e}")

    def _start(self):
        if self.is_running:
            return

        greffe = self.var_greffe.get().strip()
        if greffe not in GREFFE_MAP:
            messagebox.showerror("Erreur", f"Greffe inconnu : {greffe!r}")
            return
        try:
            annee = int(self.var_annee.get().strip())
        except ValueError:
            messagebox.showerror("Erreur", "L'année doit être un nombre entier.")
            return
        forme = self.var_forme.get().strip()
        if not forme:
            messagebox.showerror("Erreur", "La forme juridique est obligatoire.")
            return

        type_label = self.var_type.get()
        type_code = next((k for k, v in TYPE_LABELS.items() if v == type_label), "B")

        out_dir_str = self.var_output.get().strip() or str(Path.home() / "Desktop")
        out_dir = Path(out_dir_str)
        try:
            out_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            messagebox.showerror("Erreur", f"Dossier invalide :\n{e}")
            return
        output_file = out_dir / f"rcs_{greffe.replace(' ', '_')}_{annee}_{forme}.xlsx"
        self.output_path = str(output_file)

        self.is_running = True
        self.btn_run.config(state="disabled", text="Extraction en cours…")
        self.btn_open.grid_remove()
        self.progress["value"] = 0
        self.var_percent.set("0 %")
        self.var_eta.set("Initialisation…")
        self.var_status.set(f"Connexion à {BASE_URL}…")
        self.var_result.set("")
        self.lbl_result.configure(style="Success.TLabel")
        self.start_time = time.time()

        params = {
            "greffe": greffe,
            "greffe_id": GREFFE_MAP[greffe],
            "annee": str(annee),
            "forme": forme,
            "type": type_code,
            "output_file": str(output_file),
        }
        self.worker = threading.Thread(target=self._run_worker, args=(params,), daemon=True)
        self.worker.start()

    def _run_worker(self, params):
        try:
            def cb(event, **kwargs):
                self.queue.put((event, kwargs))

            session = make_session()
            rows = extract_entreprises(session, params["type"], params["greffe_id"],
                                       params["annee"], params["forme"],
                                       sleep_s=1.0, progress_cb=cb)
            cb("status", message="Génération du fichier Excel…")
            wb = build_workbook(rows, {
                "greffe": params["greffe"], "annee": params["annee"],
                "forme": params["forme"], "type": params["type"],
            })
            wb.save(params["output_file"])
            self.queue.put(("done", {"output": params["output_file"], "count": len(rows)}))
        except Exception as e:
            self.queue.put(("error", {"message": str(e)}))

    def _poll_queue(self):
        try:
            while True:
                event, data = self.queue.get_nowait()
                self._handle_event(event, data)
        except queue.Empty:
            pass
        self.root.after(120, self._poll_queue)

    def _handle_event(self, event, data):
        if event == "status":
            self.var_status.set(data.get("message", ""))
        elif event == "progress":
            current = data.get("current", 0)
            total = data.get("total", 0)
            pct = (current / total * 100) if total > 0 else 0
            self.progress["value"] = pct
            self.var_percent.set(f"{pct:.0f} %   ({current}/{total})")
            eta = data.get("eta", 0)
            if eta > 0:
                self.var_eta.set(f"≈ {format_duration(eta)} restant")
            else:
                self.var_eta.set("Bientôt terminé…")
            if "message" in data:
                self.var_status.set(data["message"])
        elif event == "done":
            self.is_running = False
            self.btn_run.config(state="normal", text="Lancer l'extraction")
            self.progress["value"] = 100
            self.var_percent.set("100 %")
            elapsed = time.time() - self.start_time if self.start_time else 0
            self.var_eta.set(f"Terminé en {format_duration(elapsed)}")
            self.var_status.set("Extraction terminée avec succès.")
            self.lbl_result.configure(style="Success.TLabel")
            self.var_result.set(
                f"✓ {data['count']} entreprise(s) extraite(s)\n"
                f"Fichier : {data['output']}"
            )
            self.btn_open.grid(row=5, column=0, columnspan=2, sticky="w", pady=(12, 0))
            messagebox.showinfo(
                "Extraction terminée",
                f"{data['count']} entreprise(s) extraite(s) en {format_duration(elapsed)}.\n\n"
                f"Fichier créé :\n{data['output']}"
            )
        elif event == "error":
            self.is_running = False
            self.btn_run.config(state="normal", text="Lancer l'extraction")
            self.var_eta.set("")
            self.var_status.set("L'extraction a échoué.")
            self.lbl_result.configure(style="Error.TLabel")
            self.var_result.set(f"✗ Erreur : {data['message']}")
            messagebox.showerror("Erreur", f"L'extraction a échoué :\n\n{data['message']}")


def launch_gui():
    if tk is None:
        return False
    try:
        root = tk.Tk()
    except Exception:
        return False
    ScraperGUI(root)
    root.mainloop()
    return True


def _auto_launch_ui():
    """Choisit automatiquement l'interface graphique ou le mode interactif."""
    if tk is None:
        if getattr(sys, "frozen", False):
            return
        launch_interactive(reason=f"Tkinter n'est pas installé ({_TK_IMPORT_ERROR}).")
        return
    if can_use_gui():
        if launch_gui():
            return
        if getattr(sys, "frozen", False):
            return
    launch_interactive(reason="L'interface graphique Tkinter n'est pas compatible avec votre installation Python.")


def main():
    if len(sys.argv) == 1:
        _auto_launch_ui()
        return

    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--greffe", default="Nosy Be", help="Nom du greffe (défaut: Nosy Be)")
    parser.add_argument("--annee", default="2025", help="Année (défaut: 2025)")
    parser.add_argument("--forme", default="SARL", help="Code forme juridique (défaut: SARL)")
    parser.add_argument("--type", default="B", choices=list(TYPE_LABELS.keys()), help="Type d'assujetti (défaut: B = Personne morale)")
    parser.add_argument("--out", default=None, help="Fichier Excel de sortie (défaut: rcs_<greffe>_<annee>_<forme>.xlsx)")
    parser.add_argument("--sleep", type=float, default=1.0, help="Pause entre requêtes en secondes (défaut: 1.0)")
    parser.add_argument("--limit", type=int, default=0, help="Limiter à N entreprises (0 = tout, utile pour tester)")
    parser.add_argument("--interactive", "-i", action="store_true", help="Forcer le mode interactif texte (pas d'interface graphique)")
    parser.add_argument("--gui", action="store_true", help="Forcer l'interface graphique (échec si indisponible)")
    args = parser.parse_args()

    if args.interactive:
        launch_interactive()
        return
    if args.gui:
        if not launch_gui():
            print("Erreur : l'interface graphique n'est pas disponible sur ce système.", file=sys.stderr)
            sys.exit(1)
        return

    if args.greffe not in GREFFE_MAP:
        print(f"Greffe inconnu: {args.greffe!r}", file=sys.stderr)
        print(f"Greffes disponibles: {', '.join(sorted(GREFFE_MAP))}", file=sys.stderr)
        sys.exit(1)

    greffe_id = GREFFE_MAP[args.greffe]
    if args.out:
        out_path = args.out
    else:
        desktop = Path.home() / "Desktop"
        desktop.mkdir(exist_ok=True)
        out_path = str(desktop / f"rcs_{args.greffe.replace(' ', '_')}_{args.annee}_{args.forme}.xlsx")

    session = make_session()

    print(f"-> Recherche: greffe={args.greffe} (id={greffe_id}) | année={args.annee} | forme={args.forme} | type={args.type}")
    rows = extract_entreprises(session, args.type, greffe_id, args.annee, args.forme,
                               sleep_s=args.sleep, limit=args.limit)

    if not rows:
        print("-> Aucun résultat. Le fichier Excel sera créé vide.")

    params = {
        "greffe": args.greffe,
        "annee": args.annee,
        "forme": args.forme,
        "type": args.type,
    }
    wb = build_workbook(rows, params)
    wb.save(out_path)
    print(f"\n-> Fichier écrit: {out_path}")
    print(f"-> {len(rows)} ligne(s) dans la feuille '{args.greffe[:31]}'")


if __name__ == "__main__":
    main()
